"""
This module provides various functions to work with XLSX file.

Reference: https://gist.github.com/vst/269aceb5a54de0adf2cb23e3482895f8
"""

__all__ = ["read_workbook_data", "read_worksheet_data"]

from typing import Any, Dict, Iterable

from openpyxl import load_workbook  # type: ignore
from openpyxl.cell.cell import Cell  # type: ignore
from openpyxl.worksheet.worksheet import Worksheet  # type: ignore


def read_workbook_data(filepath: str) -> Iterable[Dict[str, Any]]:
    """
    Attempts to read data from the workbook at the given ``filepath`` as an iterable of :py:class:`dict` instances.

    The data is read from the very first worksheet.

    :param filepath: Path to the workbook.
    :return: an iterable of :py:class:`dict` instances.
    """
    ## Load the workbook:
    workbook = load_workbook(filepath, data_only=True, read_only=True, keep_links=False, keep_vba=False)

    ## Read the first worksheet and return:
    return read_worksheet_data(workbook.worksheets[0])


def read_worksheet_data(worksheet: Worksheet) -> Iterable[Dict[str, Any]]:
    """
    Attempts to read data from the given :py:class:`Worksheet` as an iterable of :py:class:`dict` instances.

    :param sheet: :py:class:`Worksheet` instance to read data from.
    :return: an iterable of :py:class:`dict` instances.
    """
    ## If we do not have any data rows, return right here:
    if worksheet.max_row < 2:
        return iter([])

    ## Get the rows iterator:
    rows = worksheet.rows

    ## Get the header:
    header = []
    for colno, cell in enumerate(next(rows), start=1):
        ## Attempt to get colname:
        colname = str(cell.value).strip() if cell.value else f"Col{colno}"

        ## If the colname already exists in the header, append stuff:
        while colname in header:
            colname = f"{colname}_1"

        ## Add to header:
        header.append(colname)

    ## Iterate over data rows and yield data values:
    for row in rows:
        yield dict(zip(header, (_cast(i) for i in row)))


def _cast(x: Cell) -> Any:
    """
    Returns the cell value by performing required casts.

    In particular, we want to distinguish in between ``date`` and ``datetime``
    object. It is a bit ugly, but works fine (so far).
    """
    if x.is_date and "hh:" not in x.number_format.lower():
        return x.value.date()
    return x.value


if __name__ == "__main__":
    import json
    import sys

    json.dump(list(read_workbook_data(sys.argv[1])), fp=sys.stdout, indent=2, default=str)
